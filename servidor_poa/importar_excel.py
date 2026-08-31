# -*- coding: utf-8 -*-
"""Carga a la plataforma las actividades de un POA trimestral en Excel.

    .venv\\Scripts\\python.exe servidor_poa\\importar_excel.py "C:\\ruta\\al\\archivo.xlsx"

Las fotos del Excel son "imagen en celda" de Excel 365: la celda sólo guarda #VALUE! y
la imagen vive en xl/media, unida a la celda por una cadena de referencias
(vm -> metadata -> richValue -> rel -> media) que openpyxl no lee. Aquí se recorre a mano.

Se puede correr dos veces sin duplicar: si ya existe una actividad con el mismo título
y año, se salta.
"""
from __future__ import annotations

import re
import sys
import zipfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from openpyxl import load_workbook  # noqa: E402

from app import consolidado, db, fotos as fotos_mod  # noqa: E402

HOJA = "Actividades"
PRIMERA_FILA = 3
ANIO = 2026

COL = {
    "titulo": 1, "poa": 2, "prog_nac": 8, "planeado_anual": 9,
    "plan_t1": 10, "plan_t2": 11, "plan_t3": 12, "plan_t4": 13,
    "inf_t1": 14, "inf_t2": 15, "inf_t3": 16, "inf_t4": 17,
    "planeacion": 19, "observaciones": 20, "resumen": 21, "fechas": 22,
    "participa": 23, "responsable": 24,
}

# El Excel abrevia los nombres. La forma canónica es la del directorio de la plataforma.
# Ojo con «Claudia Garcia»: su apellido real es Gracia Solís (confirmado por la Sección).
ALIAS = {
    "carlos galvez": "Carlos Alberto Gálvez Valencia",
    "karla martinez lopez": "Karla Martínez López",
    "claudia garcia": "Claudia A. Gracia Solís",
    "claudia gracia": "Claudia A. Gracia Solís",
    "diana arano": "Diana Elizabeth Arano Recio",
    "natalia hernandez": "Natalia Hernández Tangarife",
    "claudia ocampo": "Claudia Angélica Ocampo Flores",
}


def mapa_fotos(src: Path) -> dict[str, str]:
    """Devuelve {'Z3': 'xl/media/image1.png', ...} siguiendo la cadena de richData."""
    z = zipfile.ZipFile(src)
    if "xl/richData/rdrichvalue.xml" not in z.namelist():
        return {}
    meta = z.read("xl/metadata.xml").decode("utf-8")
    vm_a_future = [int(v) for v in
                   re.findall(r'<rc t="1" v="(\d+)"/>', meta[meta.find("<valueMetadata"):])]
    future_a_rv = [int(i) for i in
                   re.findall(r'<xlrd:rvb i="(\d+)"/>', meta[meta.find("<futureMetadata"):])]
    rv_a_rel = [int(m.group(1)) for m in re.finditer(
        r'<rv s="\d+"><v>(\d+)</v>', z.read("xl/richData/rdrichvalue.xml").decode("utf-8"))]
    rel_ids = re.findall(r'<rel r:id="(rId\d+)"/>',
                         z.read("xl/richData/richValueRel.xml").decode("utf-8"))
    rels = dict(re.findall(
        r'Id="(rId\d+)"[^>]*Target="\.\./([^"]+)"',
        z.read("xl/richData/_rels/richValueRel.xml.rels").decode("utf-8")))

    hoja = z.read("xl/worksheets/sheet1.xml").decode("utf-8")
    salida = {}
    for celda, vm in re.findall(r'<c r="([A-Z]+\d+)"[^>]*vm="(\d+)"', hoja):
        try:
            rv = future_a_rv[vm_a_future[int(vm) - 1]]
            salida[celda] = "xl/" + rels[rel_ids[rv_a_rel[rv]]]
        except (IndexError, KeyError):
            pass
    return salida


def limpiar(v) -> str:
    return "" if v is None else str(v).strip()


def numero(v) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def buscar_catalogo(con, poa: str) -> tuple[int | None, str]:
    """Encuentra la Actividad POA en el catálogo.

    Además del texto exacto acepta un prefijo, siempre que apunte a UNA sola actividad:
    en el Excel de origen hay filas con el texto cortado a media frase, y ahí el VLOOKUP
    devolvía vacío sin avisar. Si el prefijo es ambiguo, mejor omitir que adivinar.
    """
    fila = con.execute("SELECT id FROM catalogo_poa WHERE actividad_poa = ?", (poa,)).fetchone()
    if fila:
        return fila["id"], ""
    clave = db.norm(poa)
    if len(clave) < 20:
        return None, ""
    candidatos = [f for f in con.execute("SELECT id, actividad_poa FROM catalogo_poa")
                  if db.norm(f["actividad_poa"]).startswith(clave)]
    if len(candidatos) == 1:
        return candidatos[0]["id"], candidatos[0]["actividad_poa"]
    return None, ""


def buscar_usuario(con, nombre: str) -> tuple[int | None, str]:
    """Del nombre abreviado del Excel al usuario del directorio."""
    limpio = db.norm(nombre)
    if not limpio:
        return None, ""
    canonico = ALIAS.get(limpio, nombre)
    fila = con.execute("SELECT id, nombre FROM usuarios WHERE nombre = ?",
                       (canonico,)).fetchone()
    if fila:
        return fila["id"], fila["nombre"]
    # Sin alias: se acepta sólo si el nombre del Excel está contenido en uno del directorio
    # y no hay ambigüedad. Adivinar a medias sería peor que avisar.
    candidatos = [f for f in con.execute("SELECT id, nombre FROM usuarios")
                  if limpio and limpio in db.norm(f["nombre"])]
    if len(candidatos) == 1:
        return candidatos[0]["id"], candidatos[0]["nombre"]
    return None, ""


def main() -> None:
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    src = Path(sys.argv[1])
    if not src.exists():
        sys.exit(f"No encuentro el archivo:\n  {src}")

    con = db.conectar()
    db.crear_esquema(con)

    fotos_celda = mapa_fotos(src)
    z = zipfile.ZipFile(src)
    wb = load_workbook(src, data_only=True)
    ws = wb[HOJA]

    # Los programas nacionales vienen con «Corservación» en el Excel de origen.
    programas = {db.norm(p.replace("Conservación", "Corservacion")): p
                 for p in db.PROGRAMAS_NACIONALES}
    programas.update({db.norm(p): p for p in db.PROGRAMAS_NACIONALES})

    creadas = saltadas = fotos_ok = 0
    avisos: list[str] = []

    for r in range(PRIMERA_FILA, ws.max_row + 1):
        titulo = limpiar(ws.cell(r, COL["titulo"]).value)
        poa = limpiar(ws.cell(r, COL["poa"]).value)
        if not titulo or not poa:
            continue

        cat_id, completado = buscar_catalogo(con, poa)
        if cat_id is None:
            avisos.append(f"Fila {r}: la Actividad POA no está en el catálogo, se omitió "
                          f"«{titulo[:50]}»")
            continue
        if completado:
            avisos.append(f"Fila {r}: la Actividad POA venía cortada en el Excel («…{poa[-30:]}») "
                          f"y por eso su VLOOKUP estaba vacío. Se completó a «…{completado[-45:]}»")

        ya = con.execute("SELECT id FROM actividades WHERE titulo_norm = ? AND anio = ?",
                         (db.norm(titulo), ANIO)).fetchone()
        if ya:
            saltadas += 1
            continue

        autor_id, autor_nombre = buscar_usuario(con, limpiar(ws.cell(r, COL["participa"]).value))
        if autor_id is None:
            avisos.append(f"Fila {r}: no identifiqué a «{limpiar(ws.cell(r, COL['participa']).value)}»"
                          f" (personal que participa), se omitió la fila")
            continue
        resp_id, _ = buscar_usuario(con, limpiar(ws.cell(r, COL["responsable"]).value))
        if resp_id is None and limpiar(ws.cell(r, COL["responsable"]).value):
            avisos.append(f"Fila {r}: no identifiqué al responsable "
                          f"«{limpiar(ws.cell(r, COL['responsable']).value)}», quedó sin asignar")

        prog_excel = limpiar(ws.cell(r, COL["prog_nac"]).value)
        prog = programas.get(db.norm(prog_excel), "Ninguno")

        # El Excel reparte lo planeado y lo informado en 4+4 columnas. La plataforma
        # maneja un solo trimestre por actividad: manda el trimestre en que se ejecutó,
        # y lo planeado se concentra ahí (era común planear en uno y ejecutar en otro).
        planes = {n: numero(ws.cell(r, COL[f"plan_t{n}"]).value) for n in (1, 2, 3, 4)}
        informes = {n: numero(ws.cell(r, COL[f"inf_t{n}"]).value) for n in (1, 2, 3, 4)}
        con_informe = [n for n in (1, 2, 3, 4) if informes[n] > 0]
        con_plan = [n for n in (1, 2, 3, 4) if planes[n] > 0]
        trimestre = (con_informe or con_plan or [0])[0]
        if not trimestre:
            avisos.append(f"Fila {r}: sin trimestre (no hay planeado ni informado), se omitió")
            continue
        if len(con_informe) > 1:
            avisos.append(f"Fila {r}: informa en más de un trimestre {con_informe}; "
                          f"se tomó el {trimestre}º y se sumó el resto")

        datos = {
            "titulo": titulo,
            "catalogo_id": cat_id,
            "zona": "",          # el Excel no tiene columna de zona; se llena en la plataforma
            "programa_nacional": prog,
            "anio": ANIO,
            "trimestre": trimestre,
            "planeado": sum(planes.values()) or numero(ws.cell(r, COL["planeado_anual"]).value),
            "realizado": sum(informes.values()),
            "planeacion": "No" if db.norm(limpiar(ws.cell(r, COL["planeacion"]).value)) == "no" else "Si",
            "observaciones": limpiar(ws.cell(r, COL["observaciones"]).value),
            "fechas_ejecucion": limpiar(ws.cell(r, COL["fechas"]).value),
            "responsable_id": resp_id,
        }
        consolidado.expandir_periodo(datos)
        act_id = consolidado.insertar_actividad(con, datos, autor_id)
        parte_id = consolidado.sumar_participante(
            con, act_id, autor_id, limpiar(ws.cell(r, COL["resumen"]).value))
        creadas += 1

        orden = 0
        for col in ("Z", "AA", "AB", "AC"):
            ruta = fotos_celda.get(f"{col}{r}")
            if not ruta:
                continue
            try:
                meta = fotos_mod.procesar(z.read(ruta), Path(ruta).name)
            except (fotos_mod.FotoInvalida, KeyError) as exc:
                avisos.append(f"Fila {r}, {col}: {exc}")
                continue
            con.execute(
                """INSERT INTO fotos (participacion_id, archivo, archivo_pdf,
                                      nombre_original, bytes, bytes_pdf, bytes_original,
                                      ancho, alto, orden, creada_en)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (parte_id, meta["archivo"], meta["archivo_pdf"], meta["nombre_original"],
                 meta["bytes"], meta["bytes_pdf"], meta["bytes_original"],
                 meta["ancho"], meta["alto"], orden, db.ahora()),
            )
            orden += 1
            fotos_ok += 1
        con.commit()

    print(f"\nActividades creadas : {creadas}")
    if saltadas:
        print(f"Ya existían (saltadas): {saltadas}")
    print(f"Fotos importadas    : {fotos_ok}")
    if avisos:
        print("\nRevisar:")
        for a in avisos:
            print(f"  - {a}")
    print("\nLa columna Zona / Sitio quedó vacía: el Excel no la tiene. "
          "Se llena desde la plataforma, en cada actividad.")
    con.close()


if __name__ == "__main__":
    main()
