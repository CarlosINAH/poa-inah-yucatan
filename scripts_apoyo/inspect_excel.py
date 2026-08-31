from pathlib import Path
import json
from openpyxl import load_workbook

src = Path(r"C:\Users\Soporte\Desktop\ejemplo.xlsx")
out_dir = Path("work/excel_analysis")
out_dir.mkdir(parents=True, exist_ok=True)

wb_formula = load_workbook(src, data_only=False)
wb_values = load_workbook(src, data_only=True)

def safe(value):
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)

def cell_value(ws_values, coord):
    return ws_values[coord].value

summary = {
    "sheets": [],
    "defined_names": [],
    "data_validations": [],
    "formulas": [],
    "tables": [],
}

for ws in wb_formula.worksheets:
    wsv = wb_values[ws.title]
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), min_col=1, max_col=min(ws.max_column, 18)):
        rows.append([safe(cell.value) for cell in row])
    summary["sheets"].append({
        "title": ws.title,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "first_rows": rows,
    })
    if ws.tables:
        for table in ws.tables.values():
            summary["tables"].append({
                "sheet": ws.title,
                "name": table.name,
                "ref": table.ref,
            })
    for dv in ws.data_validations.dataValidation:
        summary["data_validations"].append({
            "sheet": ws.title,
            "type": dv.type,
            "formula1": dv.formula1,
            "formula2": dv.formula2,
            "allow_blank": dv.allowBlank,
            "ranges": str(dv.sqref),
        })
    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == "f":
                summary["formulas"].append({
                    "sheet": ws.title,
                    "cell": cell.coordinate,
                    "formula": safe(cell.value),
                    "cached_value": safe(cell_value(wsv, cell.coordinate)),
                })
                if len(summary["formulas"]) >= 300:
                    break
        if len(summary["formulas"]) >= 300:
            break

for dn in wb_formula.defined_names.values():
    destinations = []
    try:
        for title, coord in dn.destinations:
            destinations.append({"sheet": title, "coord": coord})
    except Exception:
        pass
    summary["defined_names"].append({
        "name": dn.name,
        "attr_text": dn.attr_text,
        "destinations": destinations,
    })

(out_dir / "summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

print("SHEETS")
for s in summary["sheets"]:
    print(f"- {s['title']}: {s['max_row']} rows x {s['max_column']} cols")
print("\nTABLES")
for t in summary["tables"]:
    print(f"- {t['sheet']} / {t['name']}: {t['ref']}")
print("\nDATA VALIDATIONS")
for dv in summary["data_validations"][:80]:
    print(f"- {dv['sheet']} {dv['ranges']} -> {dv['type']} {dv['formula1']}")
print("\nDEFINED NAMES")
for dn in summary["defined_names"][:80]:
    print(f"- {dn['name']}: {dn['attr_text']}")
print("\nFORMULAS SAMPLE")
for f in summary["formulas"][:80]:
    print(f"- {f['sheet']}!{f['cell']}: {f['formula']} => {f['cached_value']}")
