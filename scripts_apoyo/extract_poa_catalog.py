from pathlib import Path
import json
from openpyxl import load_workbook

src = Path(r"C:\Users\Soporte\Desktop\ejemplo.xlsx")
out_dir = Path("work/excel_analysis")
out_dir.mkdir(parents=True, exist_ok=True)

def safe(value):
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)

wb = load_workbook(src, data_only=True)
catalog_ws = wb["Hoja1"]

headers = [safe(catalog_ws.cell(row=1, column=col).value) for col in range(1, 7)]
rows = []
for row in range(2, 69):
    values = [safe(catalog_ws.cell(row=row, column=col).value) for col in range(1, 7)]
    if any(v not in (None, "") for v in values):
        rows.append(dict(zip(headers, values)))

activity_ws = wb["Actividades"]
activity_headers = [safe(activity_ws.cell(row=2, column=col).value) for col in range(1, 27)]
activity_first = []
for row in range(3, 25):
    values = [safe(activity_ws.cell(row=row, column=col).value) for col in range(1, 27)]
    if any(v not in (None, "") for v in values):
        activity_first.append(dict(zip(activity_headers, values)))

payload = {
    "catalog_headers": headers,
    "activities_count": len(rows),
    "activities": rows,
    "activity_sheet_headers": activity_headers,
    "activity_sample": activity_first,
}

(out_dir / "poa_catalog.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

print("CATALOG HEADERS")
for i, h in enumerate(headers, start=1):
    print(f"{i}. {h}")

print(f"\nACTIVITIES: {len(rows)}")
for r in rows[:20]:
    print(" | ".join(str(r.get(h, "")) for h in headers))

print("\nACTIVIDADES HEADERS")
for i, h in enumerate(activity_headers, start=1):
    print(f"{i}. {h}")

print("\nSAMPLE CAPTURED ROWS")
for r in activity_first[:5]:
    print(json.dumps(r, ensure_ascii=False))
